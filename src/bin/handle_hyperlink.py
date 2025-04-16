def split_hyperlink_column(excel_file_path):
    from openpyxl import load_workbook
            
    wb = load_workbook(excel_file_path)
    ws = wb.active 
    max_col = ws.max_column  # 获取当前工作表的最大列数
    new_col_letter = chr(65 + max_col)  # 计算新列的字母标识，65是'A'的ASCII码
            
    # 添加新列的头部（假设列名为"Hyperlink"）
    new_col_head_cell = ws[f"{new_col_letter}1"]
    new_col_head_cell.value = "Hyperlink"
            
    # 遍历第一列（假设第一列是"A"列）
    for row_num in range(2, ws.max_row + 1):  # 从第二行开始，假设第一行是标题行
        cell = ws[f"A{row_num}"]
        new_cell = ws[f"{new_col_letter}{row_num}"]

        if cell.hyperlink:
            new_cell.value = cell.hyperlink.target  # 复制超链接到新列
            
    wb.save(excel_file_path)

if __name__ == '__main__':
    # get path as argument
    import sys
    excel_file_path = sys.argv[1]
    split_hyperlink_column(excel_file_path)